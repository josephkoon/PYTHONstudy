import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string

#creates workbook file
wb = openpyxl.load_workbook('example.xlsx')

#create a workbook object or use get_active_sheet() 
sheet = wb.get_sheet_by_name('sheet 1')

print(sheet['A1'])

#get value from cell A1
print(sheet['A1'].value)

#get value from cell B1
c = sheet['B1']
print(c.value)

#return location of a cell
print(c.coordinate)


#you can also use the row and column to get values
sheet.cell(row=1, column=2)

sheet.cell(row=1, column=2).value

#print out a whole column of values
for i in range(1,8,2):
	print(i,sheet.cell(row=i, column=1).value)
	
	

#convert from column number
print(get_column_letter(1))

print(column_index_from_string('A'))


#slice a sheet by number in an area
print(tuple(sheet['A1':'C3']))

#read along rows by slicing it in an area.
for rows in sheet['A1':'C3']:
	for cell in rows:
		print(cell.coordinate, cell.value)
		
#read along columns
print(sheet.columns[1])

for cell in sheet.columns[1]:
	print(cell.value)
	print('end of row!')




