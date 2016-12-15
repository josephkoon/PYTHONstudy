import openpyxl

#open and read the cells of an excel document
wb = openpyxl.load_workbook('censuspopdata.xlsx')

sheet = wb.get_sheet_by_name('Population by Census Tract')


#calculate all the tract and population data and store it in a data structure
total = 0



for i in range(1,1000):
	if str(sheet.cell(row=i, column=2).value) == 'AL':
		total+= int(sheet.cell(row=i, column=4).value)
	
print(total)




#write the data structure to a text file 


